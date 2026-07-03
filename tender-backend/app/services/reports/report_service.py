"""Report generation service for PDF and Excel reports."""

import logging
from datetime import datetime
from io import BytesIO
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from app.config import settings

logger = logging.getLogger(__name__)

HEADER_FILL = PatternFill(start_color="1a73e8", end_color="1a73e8", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
CELL_FONT = Font(size=10)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


class ReportService:
    """Generates PDF and Excel reports."""

    # ── PDF ──────────────────────────────────────────────────────────

    async def generate_pdf_report(
        self,
        tenders: list[dict],
        company_name: str,
        title: str = "Tender Report",
    ) -> bytes:
        """Generate a PDF report for matched tenders."""
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=20 * mm, bottomMargin=20 * mm)
        styles = getSampleStyleSheet()

        elements = []

        title_style = ParagraphStyle(
            "CustomTitle", parent=styles["Title"], fontSize=18, spaceAfter=12
        )
        elements.append(Paragraph(title, title_style))
        elements.append(Paragraph(f"Company: {company_name}", styles["Normal"]))
        elements.append(
            Paragraph(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}", styles["Normal"])
        )
        elements.append(Spacer(1, 12))

        if tenders:
            table_data = [["#", "Title", "Category", "Amount", "Score"]]
            for i, t in enumerate(tenders[:50], 1):
                table_data.append(
                    [
                        str(i),
                        str(t.get("title", ""))[:60],
                        str(t.get("category", "")),
                        f"{t.get('amount', 0):,.0f} UZS",
                        f"{t.get('score', 0):.0f}%",
                    ]
                )

            table = Table(table_data, colWidths=[15 * mm, 80 * mm, 30 * mm, 30 * mm, 20 * mm])
            table.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1a73e8")),
                        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                        ("FONTSIZE", (0, 0), (-1, -1), 8),
                        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
                    ]
                )
            )
            elements.append(table)

        doc.build(elements)
        return buffer.getvalue()

    async def generate_applications_pdf(
        self,
        applications: list[dict],
        company_name: str,
    ) -> bytes:
        """Generate PDF report for tender applications/pipeline."""
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=20 * mm, bottomMargin=20 * mm)
        styles = getSampleStyleSheet()

        elements = []
        title_style = ParagraphStyle(
            "CustomTitle", parent=styles["Title"], fontSize=18, spaceAfter=12
        )
        elements.append(Paragraph("Arizalar hisoboti", title_style))
        elements.append(Paragraph(f"Company: {company_name}", styles["Normal"]))
        elements.append(
            Paragraph(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}", styles["Normal"])
        )
        elements.append(Spacer(1, 12))

        if applications:
            table_data = [["#", "Tender", "Bosqich", "Narx", "Ehtimollik", "Natija"]]
            for i, a in enumerate(applications[:50], 1):
                table_data.append(
                    [
                        str(i),
                        str(a.get("tender_title", ""))[:50],
                        str(a.get("stage", "")),
                        f"{a.get('bid_amount', 0):,.0f} UZS" if a.get("bid_amount") else "-",
                        f"{a.get('win_probability', 0):.0f}%" if a.get("win_probability") else "-",
                        str(a.get("result", "-") or "-"),
                    ]
                )

            table = Table(
                table_data, colWidths=[12 * mm, 65 * mm, 25 * mm, 30 * mm, 25 * mm, 20 * mm]
            )
            table.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1a73e8")),
                        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                        ("FONTSIZE", (0, 0), (-1, -1), 8),
                        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
                    ]
                )
            )
            elements.append(table)

        doc.build(elements)
        return buffer.getvalue()

    async def generate_analytics_pdf(
        self,
        stats: dict,
        company_name: str,
    ) -> bytes:
        """Generate PDF report for analytics summary."""
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=20 * mm, bottomMargin=20 * mm)
        styles = getSampleStyleSheet()

        elements = []
        title_style = ParagraphStyle(
            "CustomTitle", parent=styles["Title"], fontSize=18, spaceAfter=12
        )
        elements.append(Paragraph("Analitika hisoboti", title_style))
        elements.append(Paragraph(f"Company: {company_name}", styles["Normal"]))
        elements.append(
            Paragraph(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}", styles["Normal"])
        )
        elements.append(Spacer(1, 20))

        summary_data = [
            ["Ko'rsatkich", "Qiymat"],
            ["Jami tenderlar", str(stats.get("total_tenders", 0))],
            ["Faol tenderlar", str(stats.get("active_tenders", 0))],
            ["Yuborilgan arizalar", str(stats.get("total_applications", 0))],
            ["Yutilgan tenderlar", str(stats.get("won_tenders", 0))],
            ["Yutish foizi", f"{stats.get('win_rate', 0):.1f}%"],
            ["Umumiy summa", f"{stats.get('total_amount', 0):,.0f} UZS"],
        ]
        table = Table(summary_data, colWidths=[80 * mm, 80 * mm])
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1a73e8")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTSIZE", (0, 0), (-1, -1), 10),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
                ]
            )
        )
        elements.append(table)

        doc.build(elements)
        return buffer.getvalue()

    # ── Excel ────────────────────────────────────────────────────────

    async def generate_tenders_excel(
        self,
        tenders: list[dict],
        company_name: str,
    ) -> bytes:
        """Generate Excel report for tenders."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Tenderlar"

        headers = ["#", "Nomi", "Kategoriya", "Region", "Summa", "Status", "Deadline", "Bal"]
        self._write_excel_header(ws, headers)

        for i, t in enumerate(tenders, 1):
            row = i + 1
            ws.cell(row=row, column=1, value=i)
            ws.cell(row=row, column=2, value=str(t.get("title", ""))[:80])
            ws.cell(row=row, column=3, value=str(t.get("category", "")))
            ws.cell(row=row, column=4, value=str(t.get("region", "")))
            ws.cell(row=row, column=5, value=t.get("amount", 0))
            ws.cell(row=row, column=6, value=str(t.get("status", "")))
            ws.cell(row=row, column=7, value=str(t.get("deadline", "")))
            ws.cell(row=row, column=8, value=t.get("score", 0))
            self._style_excel_row(ws, row, len(headers))

        ws.column_dimensions["B"].width = 50
        ws.column_dimensions["E"].number_format = "#,##0"

        self._add_meta_sheet(wb, company_name, "Tenderlar hisoboti")
        return self._save_workbook(wb)

    async def generate_applications_excel(
        self,
        applications: list[dict],
        company_name: str,
    ) -> bytes:
        """Generate Excel report for applications pipeline."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Arizalar"

        headers = ["#", "Tender", "Bosqich", "Prioritet", "Narx", "Ehtimollik", "Natija", "Sana"]
        self._write_excel_header(ws, headers)

        for i, a in enumerate(applications, 1):
            row = i + 1
            ws.cell(row=row, column=1, value=i)
            ws.cell(row=row, column=2, value=str(a.get("tender_title", ""))[:60])
            ws.cell(row=row, column=3, value=str(a.get("stage", "")))
            ws.cell(row=row, column=4, value=str(a.get("priority", "")))
            ws.cell(row=row, column=5, value=a.get("bid_amount", 0) or 0)
            ws.cell(row=row, column=6, value=a.get("win_probability", 0) or 0)
            ws.cell(row=row, column=7, value=str(a.get("result", "") or ""))
            ws.cell(row=row, column=8, value=str(a.get("submitted_at", "") or ""))
            self._style_excel_row(ws, row, len(headers))

        ws.column_dimensions["B"].width = 40

        self._add_meta_sheet(wb, company_name, "Arizalar hisoboti")
        return self._save_workbook(wb)

    async def generate_analytics_excel(
        self,
        stats: dict,
        tenders_by_category: list[dict],
        tenders_by_region: list[dict],
        company_name: str,
    ) -> bytes:
        """Generate Excel report for analytics with multiple sheets."""
        wb = Workbook()

        # Sheet 1: Summary
        ws = wb.active
        ws.title = "Xulosa"
        summary_rows = [
            ("Ko'rsatkich", "Qiymat"),
            ("Jami tenderlar", stats.get("total_tenders", 0)),
            ("Faol tenderlar", stats.get("active_tenders", 0)),
            ("Yuborilgan arizalar", stats.get("total_applications", 0)),
            ("Yutilgan tenderlar", stats.get("won_tenders", 0)),
            ("Yutish foizi (%)", stats.get("win_rate", 0)),
            ("Umumiy summa (UZS)", stats.get("total_amount", 0)),
        ]
        self._write_excel_header(ws, list(summary_rows[0]))
        for i, (label, value) in enumerate(summary_rows[1:], 2):
            ws.cell(row=i, column=1, value=label)
            ws.cell(row=i, column=2, value=value)
            self._style_excel_row(ws, i, 2)
        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 20

        # Sheet 2: By category
        ws2 = wb.create_sheet("Kategoriya bo'yicha")
        headers2 = ["Kategoriya", "Soni", "Umumiy summa"]
        self._write_excel_header(ws2, headers2)
        for i, c in enumerate(tenders_by_category, 2):
            ws2.cell(row=i, column=1, value=str(c.get("category", "")))
            ws2.cell(row=i, column=2, value=c.get("count", 0))
            ws2.cell(row=i, column=3, value=c.get("total_amount", 0))
            self._style_excel_row(ws2, i, len(headers2))
        ws2.column_dimensions["A"].width = 25

        # Sheet 3: By region
        ws3 = wb.create_sheet("Region bo'yicha")
        headers3 = ["Region", "Soni", "Umumiy summa"]
        self._write_excel_header(ws3, headers3)
        for i, r in enumerate(tenders_by_region, 2):
            ws3.cell(row=i, column=1, value=str(r.get("region", "")))
            ws3.cell(row=i, column=2, value=r.get("count", 0))
            ws3.cell(row=i, column=3, value=r.get("total_amount", 0))
            self._style_excel_row(ws3, i, len(headers3))
        ws3.column_dimensions["A"].width = 25

        self._add_meta_sheet(wb, company_name, "Analitika hisoboti")
        return self._save_workbook(wb)

    # ── Weekly digest (existing) ─────────────────────────────────────

    async def generate_weekly_digest(
        self,
        tenders: list[dict],
        stats: dict,
    ) -> str:
        """Generate an HTML weekly digest email body."""
        tender_rows = ""
        for t in tenders[:20]:
            tender_rows += (
                f"<tr>"
                f"<td>{t.get('title', '')[:80]}</td>"
                f"<td>{t.get('category', '')}</td>"
                f"<td>{t.get('amount', 0):,.0f} UZS</td>"
                f"<td>{t.get('score', 0):.0f}%</td>"
                f"</tr>"
            )

        return f"""
        <html>
        <body style="font-family: Arial, sans-serif;">
        <h2>TenderIQ - Haftalik Xulosa</h2>
        <p>Yangi tenderlar: {stats.get('new_tenders', 0)}</p>
        <p>Mos kelgan tenderlar: {stats.get('matches', 0)}</p>
        <table border="1" cellpadding="8" cellspacing="0" style="border-collapse: collapse;">
        <tr style="background: #1a73e8; color: white;">
        <th>Tender</th><th>Kategoriya</th><th>Summa</th><th>Bal</th>
        </tr>
        {tender_rows}
        </table>
        </body>
        </html>
        """

    # ── Private helpers ──────────────────────────────────────────────

    def _write_excel_header(self, ws, headers: list[str]) -> None:
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center")
            cell.border = THIN_BORDER

    def _style_excel_row(self, ws, row: int, num_cols: int) -> None:
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = CELL_FONT
            cell.border = THIN_BORDER

    def _add_meta_sheet(self, wb: Workbook, company_name: str, report_title: str) -> None:
        ws = wb.create_sheet("Ma'lumot")
        ws.cell(row=1, column=1, value="Hisobot").font = Font(bold=True)
        ws.cell(row=1, column=2, value=report_title)
        ws.cell(row=2, column=1, value="Kompaniya").font = Font(bold=True)
        ws.cell(row=2, column=2, value=company_name)
        ws.cell(row=3, column=1, value="Yaratilgan").font = Font(bold=True)
        ws.cell(row=3, column=2, value=datetime.now().strftime("%Y-%m-%d %H:%M"))
        ws.column_dimensions["A"].width = 15
        ws.column_dimensions["B"].width = 40

    def _save_workbook(self, wb: Workbook) -> bytes:
        buffer = BytesIO()
        wb.save(buffer)
        return buffer.getvalue()
